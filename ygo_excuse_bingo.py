import random
import argparse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, PageBreak, KeepTogether
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle

# Excuses pool
excuses = [
    "My hand was unplayable",
    "I drew my brick card(s)",
    "I only drew non-engine",
    "I only drew engine",
    "You had a custom hand",
    "I didn’t see any starters",
    "I didn’t see my Side Deck cards",
    "I sided wrong",
    "I always lose against you",
    "The judge call was wrong",
    "I misplayed",
    "I forgot to activate my effect",
    "Dice roll screwed me",
    "Your deck only wins when you go first",
    "You had the 1-of",
    "You drew the out",
    "My topdecks were bad",
    "Your topdecks were crazy"
    "My deck isn’t finished",
    "I’m still learning the deck",
    "I was testing tech cards",
    "I didn’t know what your cards did",
    "Bad matchup",
    "Time rules screwed me",
    "That’s not how it works on Master Duel",
    "I would’ve won next turn",
    "I don’t know the matchup",
    "You’re playing a higher tier deck"
]

# ---------------- Board Generation ----------------
def generate_bingo_board():
    """Generate one 5x5 bingo board as a 2D list."""
    selected = random.sample(excuses, 24)
    board = selected[:12] + ["FREE SPACE"] + selected[12:]
    return [board[i:i+5] for i in range(0, 25, 5)]

def generate_multiple_boards(num_boards):
    """Generate a list of bingo boards."""
    return [generate_bingo_board() for _ in range(num_boards)]

# ---------------- Excel Export ----------------
def save_boards_to_excel(boards, filename="bingo_boards.xlsx"):
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    font = Font(size=12, bold=True)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    for i, bingo_matrix in enumerate(boards, 1):
        ws = wb.create_sheet(title=f"Board {i}")
        for r, row in enumerate(bingo_matrix, 1):
            for c, val in enumerate(row, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.font = font
                cell.alignment = alignment
                cell.border = border
                ws.column_dimensions[chr(64+c)].width = 25
            ws.row_dimensions[r].height = 60

    wb.save(filename)
    print(f"{len(boards)} bingo boards saved to {filename}")

# ---------------- PDF Export ----------------
def save_boards_to_pdf(boards, filename="bingo_boards.pdf"):
    pdf = SimpleDocTemplate(filename, pagesize=landscape(letter))
    elements = []

    cell_style = ParagraphStyle(name="ExcuseStyle", fontName="Helvetica-Bold", fontSize=10,
                                alignment=1, leading=12)
    header_style = ParagraphStyle(name="HeaderStyle", fontName="Helvetica-Bold", fontSize=16,
                                  alignment=1, spaceAfter=12)

    MIN_ROW_HEIGHT = 70
    COL_WIDTH = 120

    for i, bingo_matrix in enumerate(boards):
        wrapped_matrix = [[Paragraph(cell, cell_style) for cell in row] for row in bingo_matrix]

        table = Table(wrapped_matrix, colWidths=[COL_WIDTH]*5, rowHeights=[MIN_ROW_HEIGHT]*5)
        table.setStyle(TableStyle([
            ("GRID", (0,0), (-1,-1), 1, colors.black),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("ALIGN", (0,0), (-1,-1), "CENTER"),
            ("BACKGROUND", (0,0), (-1,-1), colors.whitesmoke),
        ]))

        elements.append(KeepTogether([Paragraph("Yu-Gi-Oh! Excuses Bingo", header_style), table]))

        if i < len(boards) - 1:
            elements.append(PageBreak())

    pdf.build(elements)
    print(f"{len(boards)} bingo boards saved to {filename}")

# ---------------- CLI ----------------
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate Yu-Gi-Oh! Excuses Bingo boards.")
    parser.add_argument("--num", type=int, default=5, help="Number of boards to generate (default: 5)")
    parser.add_argument("--pdf", action="store_true", help="Export as PDF (default if no options given)")
    parser.add_argument("--excel", action="store_true", help="Export as Excel")
    parser.add_argument("--both", action="store_true", help="Export both Excel and PDF")
    args = parser.parse_args()

    # Generate the boards once
    boards = generate_multiple_boards(args.num)

    if args.both:
        save_boards_to_excel(boards)
        save_boards_to_pdf(boards)
    elif args.excel:
        save_boards_to_excel(boards)
    else:  # default PDF-only
        save_boards_to_pdf(boards)
